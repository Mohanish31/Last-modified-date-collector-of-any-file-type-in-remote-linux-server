import openpyxl
import sys
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pysftp
from datetime import datetime
import time
import xlrd
import dateutil.parser
import logging


def readDataExcel(profileCount, filePathCount, inputFile):

    print("Started reading input file")
    profileList ,filePathList ,fileNameList,serverList,expected_Date_list = [],[],[],[],[]
    wb_obj = openpyxl.load_workbook(inputFile)
    sheet_obj = wb_obj["Server_Details"]
    for row in range(profileCount):serverList.append(sheet_obj.cell(row=row + 2, column=1).value)
    sheet_obj = wb_obj["Server_Details"]
    for row in range(profileCount):profileList.append(sheet_obj.cell(row=row + 2, column=2).value)
    sheet_obj = wb_obj["File_Details"]
    for row in range(filePathCount):filePathList.append(sheet_obj.cell(row=row + 2, column=1).value)  # collected list of profile
    sheet_obj = wb_obj["File_Details"]
    for row in range(filePathCount):fileNameList.append(sheet_obj.cell(row=row + 2, column=2).value)  # collected list of profile
    for row in range(filePathCount): expected_Date_list.append(sheet_obj.cell(row=row + 2, column=3).value)
    return serverList, profileList, filePathList, fileNameList,expected_Date_list


def countRows(inputFile,sheetindex):

    print("Started Couning rows from input file")
    book = xlrd.open_workbook(inputFile)
    sheet = book.sheet_by_index(sheetindex)      
    return (sheet.nrows - 1)


def increaseServer(serverList,profileCountExcel,filePathCountExcel):

    print("inceasing server count")
    servers = []
    for i in range(profileCountExcel):
        Servername = serverList[i]
        for j in range(filePathCountExcel):
            servers.append(Servername)
    #print("List fo server :", servers)
    return(servers)

def fullpath_of_file (profileList,filePathList,fileNameList,profileCountExcel,filePathCountExcel):

    print("Creating full path")
    fullPaths = []
    for i in range(profileCountExcel):
        for j in range(filePathCountExcel):
            fullPath = profileList[i] + filePathList[j] + fileNameList[j]
            fullPaths.append(fullPath)
    return(fullPaths)

def write(servers,fullPathlist,inputFile):
    
    print("Started writing input file")
    wb_obj = openpyxl.load_workbook(inputFile)
    sheet_obj = wb_obj["Result"]
    Srows = 0
    frows = 0
    for S in (servers):
        cell_obj = sheet_obj.cell(row=Srows + 2, column=1)
        cell_obj.value = S
        #print(S)
        Srows +=1
    for f in fullPathlist:
        cell_obj = sheet_obj.cell(row=frows + 2, column=2)
        cell_obj.value = f
        #print(f)
        frows +=1
    wb_obj.save(inputFile)

def Connect(servers, fullPathlist,filePathCountExcel,profileCountExcel):

    print("Connecting servers")
    Last_modifiedList = []
    for i in range(profileCountExcel * filePathCountExcel):
        myHostname = servers[i]
	# changes the host and password
        myUsername = "hostname"   
        myPassword = "password"
        with pysftp.Connection(host=myHostname, username=myUsername, password=myPassword) as sftp:
            try: 
                print("Successfully connected with :",myHostname )
                utime = sftp.stat(fullPathlist[i]).st_mtime
           	Last_modified = str(datetime.fromtimestamp(utime))
                #Last_modified = dateutil.parser.parse(Last_modified).date()
                print("Path"+fullPathlist[i] +""+Last_modified)
         	Last_modifiedList.append(Last_modified)
	    except Exception:
                Last_modified = "NA"
                Last_modifiedList.append(Last_modified)
    print(Last_modifiedList)
    return (Last_modifiedList)


def expected_Dates_of_file (expected_Date_list,profileCountExcel,filePathCountExcel):
   
    print("creating list of expected date")
    expected_Dates = []
    for i in range(profileCountExcel):
        for j in range(filePathCountExcel):
            expected_Date = expected_Date_list[j]
            expected_Dates.append(expected_Date)
    #print(expected_Dates)
    return(expected_Dates)



def write_modifiedtime(Last_modifiedList,inputFile):

    print("Started writing modified time in inputfile")
    wb_obj = openpyxl.load_workbook(inputFile)
    sheet_obj = wb_obj["Result"]
    Mrows = 0
    for M in Last_modifiedList:
        cell_obj = sheet_obj.cell(row=Mrows + 2, column=3)
        print(M)
        if M =="NA":
            cell_obj.value = M
            print(M)
        else:
            M = dateutil.parser.parse(M).date()
            cell_obj.value = M
            print(M)
        Mrows +=1
    wb_obj.save(inputFile)

def compareDate(inputFile,expected_Dates,filePathCountExcel,profileCountExcel ):

    print("Started compairing modified date with expected date in input file")
    wb_obj = openpyxl.load_workbook(inputFile)
    for rows in range(profileCountExcel  * filePathCountExcel):
        expected_Date= expected_Dates[rows]
        #print(expected_Date)
        sheet_obj = wb_obj["Result"]
        modified_Date = sheet_obj.cell(row=rows + 2, column=3).value
        #print(modified_Date)
        if expected_Date == modified_Date :
            sheet_obj = wb_obj["Result"]
            cell_obj = sheet_obj.cell(row=rows + 2, column=4)
            cell_obj.value = "Pass"
	    #print("expected_Date %s,modified_Date %s,REsult %s "%(expected_Date,modified_Date,cell_obj.value))
        elif expected_Date == None:
            cell_obj = sheet_obj.cell(row=rows + 2, column=4)
            cell_obj.value = "Expected date not found"
            #print("expected_Date %s,modified_Date %s,REsult %s "%(expected_Date,modified_Date,cell_obj.value))
        else:
            sheet_obj = wb_obj["Result"]
            cell_obj = sheet_obj.cell(row=rows + 2, column=4)
            cell_obj.value = "Fail"
            #print("expected_Date %s,modified_Date %s,REsult %s "%(expected_Date,modified_Date,cell_obj.value))
    wb_obj.save(inputFile)

def mail(inputFile):

    print("Started sending mail with inputfile")
    s = smtplib.SMTP('hostaddress')
    Email_Body = "Body Tittle"
    fromaddr = "From Tittle"
    toaddr = "abc@mail.com,xyz@mail.com "  ##mulitple email address can be used but sepreated by comma
    msg = MIMEMultipart()
    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Subject Tittle"
    body = Email_Body
    msg.attach(MIMEText(body, 'plain'))
    filename = 'filename.xlsx'
    attachment = open(inputFile)
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read())
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p)
    text = msg.as_string()
    # sending the mail
    for reception in toaddr.split(','):
        s.sendmail(fromaddr, reception, text)
        #print(reception)
        #print("sent mail")
        # terminating the session
    s.quit()


# Main Function
if __name__ == '__main__':

    start = time.time()  #Start time

    servers,modifiedList,serverList,profileList,filePathList,fileNameList,expected_Dates = [],[],[],[],[],[],[]
    try :
        inputFile = sys.argv[1:]
        inputFile = inputFile[0]
        #print(inputFile)
    except:logging.critical('Please check the input file path or file name')

    profileCountExcel = countRows(inputFile,sheetindex=0)   #count profile in file
    if profileCountExcel == 0:logging.critical('Profile Count is zero please check the file')
    filePathCountExcel = countRows(inputFile,sheetindex=1)  #Count FilePath in file
    if filePathCountExcel == 0:logging.critical('filepath Count is zero please check the file')

    try :serverList,profileList,filePathList,fileNameList,expected_Date_list = readDataExcel(profileCountExcel,filePathCountExcel,inputFile)
    except:print("Please put the correct path or file")

    #print(serverList,profileList,filePathList,fileNameList)

    servers = increaseServer(serverList,profileCountExcel,filePathCountExcel)
    fullPaths = fullpath_of_file(profileList,filePathList,fileNameList,profileCountExcel,filePathCountExcel)

    try :write(servers,fullPaths, inputFile)
    except Exception :logging.exception("Exception occurred in Writing Server and fullpath") 
    try :modifiedList = Connect (servers, fullPaths,filePathCountExcel,profileCountExcel)
    except Exception :logging.exception("Connection issue with server")
    try :write_modifiedtime(modifiedList, inputFile)
    except Exception :logging.exception("Exception occurred in modified list")

    
    expected_Dates = expected_Dates_of_file (expected_Date_list,profileCountExcel,filePathCountExcel)

    try :compareDate(inputFile,expected_Dates,filePathCountExcel,profileCountExcel )
    except Exception :logging.exception("Exception occurred in compareDate")

    # Disable the mail

    mail(inputFile)
    
    end = time.time() #End time
    print("Successfully executer: total time taken to execute is ", end - start)



